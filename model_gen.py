#!/usr/bin/python
# -*- coding: UTF-8 -*-
import sys, os
from openpyxl import load_workbook
from datetime import datetime
from utils import writer
from utils.model import Model, ModelField, ModelFieldType, ParseOneToOneFieldName, ParseForeignKeyName
from utils.enum import Enum, EnumField

COL_2_ROW_1 = 'B1'  # ClassName
COL_2_ROW_2 = 'B2'  # ClassDescription
COL_2_ROW_3 = 'B3'  # ClassGroup

ROW_4 = '4'         # FieldName
ROW_5 = '5'         # FieldDescription
ROW_6 = '6'         # Type
ROW_7 = '7'         # Uniqueness
ROW_8 = '8'         # Required
ROW_9 = '9'         # Server
ROW_10 = '10'       # Client

def ParseModelFieldType(field):
    if field == 'id':
        return ModelFieldType.IntegerField
    elif field == 'bool':
        return ModelFieldType.BooleanField
    elif field == 'string':
        return ModelFieldType.TextField
    elif field == 'int':
        return ModelFieldType.IntegerField
    elif field == 'float':
        return ModelFieldType.FloatField
    elif field.startswith('id('):
        return ModelFieldType.OneToOneField
    elif field.startswith('array(id('):
        return ModelFieldType.ForeignKey
    elif field.startswith('array('):
        return ModelFieldType.TextField
    elif field.startswith('enum('):
        return ModelFieldType.IntegerField
    elif field.startswith('Enum'):
        return ModelFieldType.IntegerField
    raise Exception('Unknown field type ' + field)

def ParseModelFieldUnique(field):
    if field == 'None':
        return False
    return True

def ParseModelFieldEnum(field):
    if field.startswith('enum('):
        content_parts = field[field.index('(') + 1:field.rindex(')')].split(',')
        enum_name = content_parts[0]
        enum = Enum(enum_name)
        # parse enum fields
        for i in range(1, len(content_parts)):
            content_part = content_parts[i]
            field_name = content_part[0:content_part.index('(')]
            field_description = content_part[content_part.index('(') + 1:content_part.index(')')]
            field = EnumField(field_name, i - 1, field_description)
            enum.AddField(field)
        return enum
    return None

def GenerateHeader(file):
    writer.I0(file, "# This file is auto-generated, please don't modify it directly.")
    writer.I0(file, "# Modify source xls file and use model_gen to regenerate again.")
    writer.I0(file, "#")
    writer.I0(file, "# Last generate time: " + str(datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    writer.I0(file)

def main():
    if len(sys.argv) < 2:
        print("Source file folder is required. Usage:")
        print("     python model_gen.py source_folder/ # 'source_folder' is a relative path")
        print()
        print("Target file folder is optional. Usage:")
        print("     python model_gen.py source_folder/ target_folder/ # 'target_folder' is a relative path")
        return

    SOURCE_FOLDER = os.getcwd() + '/' + sys.argv[1]
    TARGET_FOLDER = SOURCE_FOLDER
    xls_files = list()
    if len(sys.argv) >= 3:
        TARGET_FOLDER = os.getcwd() + '/' + sys.argv[2]
    print("Scanning xls file in", SOURCE_FOLDER)
    for file in os.listdir(SOURCE_FOLDER):
        if file.endswith(".xlsx") and not file.startswith("~"):
            xls_files.append(os.path.join(SOURCE_FOLDER, file))

    if not os.path.isdir(TARGET_FOLDER):
        os.makedirs(TARGET_FOLDER)

    if len(xls_files) == 0:
        print("Could not find xls file in", SOURCE_FOLDER)
        return
    print("Found xls files", xls_files)

    gen_models = list()
    gen_enums = list()
    MODEL_FILE = 'models.py'
    ENUM_FILE = 'enums.py'
    ADMIN_FILE = 'admin.py'
    SERIALIZER_FILE = 'serializers.py'

    # models.py
    model_file = open(TARGET_FOLDER + MODEL_FILE, 'w')
    GenerateHeader(model_file)
    writer.I0(model_file, "from django.db import models")
    writer.I0(model_file)

    # enums.py
    enum_file = open(TARGET_FOLDER + ENUM_FILE, 'w')
    GenerateHeader(enum_file)
    writer.I0(enum_file, "from enum import Enum")
    writer.I0(enum_file)

    # admin.py
    admin_file = open(TARGET_FOLDER + ADMIN_FILE, 'w')
    GenerateHeader(admin_file)
    writer.I0(admin_file, "from django.contrib import admin")
    writer.I0(admin_file, "from .models import *")
    writer.I0(admin_file)

    # serializers.py
    serializer_file = open(TARGET_FOLDER + SERIALIZER_FILE, 'w')
    GenerateHeader(serializer_file)
    writer.I0(serializer_file, "from rest_framework import serializers")
    writer.I0(serializer_file, "from .models import *")
    writer.I0(serializer_file)

    for xls_file in xls_files:
        print("Parsing xls file", xls_file)
        wb = load_workbook(xls_file, read_only=True)
        print("Found xls sheets", wb.sheetnames)
        # parse xls sheet
        for sheet in wb:
            print('Parsing xls sheet', sheet.title)
            # load model base info
            class_name = sheet[COL_2_ROW_1].value
            class_description = sheet[COL_2_ROW_2].value
            class_group = sheet[COL_2_ROW_3].value
            # create model instance
            model = Model(
                class_name,
                class_description,
                class_group)
            # load fields base info
            last_field_name = ''
            for i in range(1, len(sheet[ROW_4])):
                field_name = sheet[ROW_4][i].value
                # array field type
                print("Parsing field", field_name)
                if field_name == last_field_name:
                    continue
                field_description = sheet[ROW_5][i].value
                field_origin_type = sheet[ROW_6][i].value
                field_type = ParseModelFieldType(field_origin_type)
                field_enum = ParseModelFieldEnum(field_origin_type)
                if field_enum:
                    gen_enums.append(field_enum)
                field_unique = ParseModelFieldUnique(sheet[ROW_7][i].value)
                field_required = sheet[ROW_8][i].value
                field_server = sheet[ROW_9][i].value
                field_client = sheet[ROW_10][i].value
                last_field_name = field_name
                # create field instance
                field = ModelField(
                    field_type,
                    field_origin_type,
                    field_name,
                    field_description,
                    field_unique,
                    field_required,
                    field_server,
                    field_client)
                model.AddField(field)
            # save model to generate
            gen_models.append(model)

    # write models.py
    print("Writing " + MODEL_FILE + " file...")
    for model in gen_models:
        print("Generating model", model.name)
        # model comment
        writer.I0(model_file, "# Description: " + model.description)
        writer.I0(model_file, "# Group: " + model.group)
        # model define
        writer.I0(model_file, "class " + model.name + "(models.Model):")
        # model fields
        for field in model.fields:
            # model will auto define id
            if field.name.lower() == 'id':
                continue
            # model field comment
            writer.I1(model_file, "# Description: " + str(field.description))
            writer.I1(model_file, "# Type: " + str(field.origin_type))
            writer.I1(model_file, "# Unique: " + str(field.unique) + ", Required: " + str(field.required))
            writer.I1(model_file, "# Server: " + str(field.server) + ", Client: " + str(field.client))
            writer.I1(model_file, field.FieldDefine())
            writer.I0(model_file)
        # __str__
        writer.I1(model_file, "def __str__(self):")
        writer.I2(model_file, "return u'{}'".format(model.name))
        writer.I0(model_file)

    # write admin.py
    print("Writing " + ADMIN_FILE + " file...")
    for model in gen_models:
        print("Generating", model.name + "Admin")
        writer.I0(admin_file, "class " + model.name + "Admin(admin.ModelAdmin):")
        writer.I1(admin_file, "pass")
        writer.I0(admin_file)
        writer.I0(admin_file, "admin.site.register(" + model.name + ", " + model.name + "Admin)")
        writer.I0(admin_file)

    # write serializers.py
    print("Writing " + SERIALIZER_FILE + " file...")
    pending_models = list(gen_models)
    written_model_names = set()
    for model in pending_models:
        # check if model has relation field
        related_model_tup = list() # (0: field_name, 1: field_type)
        for field in model.fields:
            if field.type == ModelFieldType.OneToOneField or field.type == ModelFieldType.ForeignKey:
                if field.type == ModelFieldType.OneToOneField:
                    related_field = ParseOneToOneFieldName(field.origin_type)
                elif field.type == ModelFieldType.ForeignKey:
                    related_field = ParseForeignKeyName(field.origin_type)
                if not related_field:
                    raise Exception('Parse related field error ' + field.origin_type)
                tup = (field.name, related_field)
                related_model_tup.append(tup)
        # check if all related model is written before
        meet_all = True
        for model_tup in related_model_tup:
            # do not count if related to self
            if model_tup[1] not in written_model_names and model_tup[1] != model.name:
                meet_all = False
                break
        # can write current model
        if meet_all:
            print("Generating", model.name + "Serializer")
            writer.I0(serializer_file, "class " + model.name + "Serializer(serializers.ModelSerializer):")
            for model_tup in related_model_tup:
                # skip if is related to self
                if model_tup[1] == model.name:
                    continue
                # related field serializer should already generated
                writer.I1(serializer_file, model_tup[0] + " = " + model_tup[1] + "Serializer(read_only=True)")
            writer.I0(serializer_file)
            writer.I1(serializer_file, "class Meta:")
            writer.I2(serializer_file, "model = " + model.name)
            writer.I2(serializer_file, "fields = [")
            for field in model.fields:
                field_name = field.name
                if field_name == "Id":
                    field_name = "id"
                writer.I3(serializer_file, "'{}',".format(field_name))
            writer.I2(serializer_file, "]")
            writer.I0(serializer_file)
            # mark written model
            written_model_names.add(model.name)
        else:
            # push back current model for next visit
            pending_models.append(model)

    # write enums.py
    print("Writing " + ENUM_FILE + " file...")
    for enum in gen_enums:
        print("Generating enum", enum.name)
        writer.I0(enum_file, "class " + enum.name + "(Enum):")
        for field in enum.fields:
            writer.I1(enum_file, field.name + " = " + str(field.value) + " # " + field.description)
        writer.I0(enum_file)

    # close file
    model_file.close()
    enum_file.close()
    admin_file.close()
    serializer_file.close()

    print("Files generated in", TARGET_FOLDER)
    print("Success!")

main()
