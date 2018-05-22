#!/usr/bin/python
# -*- coding: UTF-8 -*-
import os
from openpyxl import load_workbook
from gen_utils.model import Model, ModelField, ModelFieldType
from gen_utils.enum import Enum, EnumField

COL_2_ROW_1 = 'B1'  # ClassName
COL_2_ROW_2 = 'B2'  # ClassDescription
COL_2_ROW_3 = 'B3'  # ClassGroup

ROW_4 = '4'         # FieldName
ROW_5 = '5'         # FieldDescription
ROW_6 = '6'         # Type
ROW_7 = '7'         # Uniqueness
ROW_8 = '8'         # Required
ROW_9 = '9'         # Server
ROW_10 = '10'       # Client

COL_2 = 'B'         # Id
DATA_START_ROW = 11

def ParseOneToOneFieldName(content):
    return content[content.index('(') + 1:content.index(')')]

def ParseForeignKeyName(content):
    return content[content.index('(', content.index('(') + 1) + 1:content.index(')')]

def ParseModelFieldType(field):
    if field == 'id':
        return ModelFieldType.IntegerField
    elif field == 'bool':
        return ModelFieldType.BooleanField
    elif field == 'string':
        return ModelFieldType.CharField
    elif field == 'int':
        return ModelFieldType.IntegerField
    elif field == 'float':
        return ModelFieldType.FloatField
    elif field.startswith('id('):
        return ModelFieldType.OneToOneField
    elif field.startswith('array(id('):
        return ModelFieldType.ForeignKey
    elif field.startswith('array('):
        return ModelFieldType.TextField
    elif field.startswith('enum('):
        return ModelFieldType.IntegerField
    elif field.startswith('Enum'):
        return ModelFieldType.IntegerField
    raise Exception('Unknown field type ' + field)

def ParseModelFieldUnique(field):
    if field == 'None':
        return False
    return True

def ParseModelFieldEnum(field):
    if field.startswith('enum('):
        content_parts = field[field.index('(') + 1:field.rindex(')')].split(',')
        enum_name = content_parts[0]
        enum = Enum(enum_name)
        # parse enum fields
        for i in range(1, len(content_parts)):
            content_part = content_parts[i]
            field_name = content_part[0:content_part.index('(')]
            field_description = content_part[content_part.index('(') + 1:content_part.index(')')]
            field = EnumField(field_name, i - 1, field_description)
            enum.AddField(field)
        return enum
    return None

def ParseModel(folder):
    xls_files = list()
    print("Scanning xls file in", folder)
    for file in os.listdir(folder):
        if file.endswith(".xlsx") and not file.startswith("~"):
            xls_files.append(os.path.join(folder, file))

    if len(xls_files) == 0:
        print("Could not find xls file in", folder)
        return False
    print("Found xls files", xls_files)

    gen_models = list()
    gen_enums = list()

    for xls_file in xls_files:
        print("Parsing xls file", xls_file)
        wb = load_workbook(xls_file, read_only=True)
        print("Found xls sheets", wb.sheetnames)
        # parse xls sheet
        for sheet in wb:
            print('Parsing xls sheet', sheet.title)
            # load model base info
            class_name = sheet[COL_2_ROW_1].value
            class_description = sheet[COL_2_ROW_2].value
            class_group = sheet[COL_2_ROW_3].value
            # create model instance
            model = Model(
                class_name,
                class_description,
                class_group,
                xls_file,
                sheet.title)
            # load fields base info
            last_field_name = ''
            for i in range(1, len(sheet[ROW_4])):
                field_name = sheet[ROW_4][i].value
                # array field type
                print("Parsing field", field_name)
                if field_name == last_field_name:
                    continue
                field_description = sheet[ROW_5][i].value
                field_origin_type = sheet[ROW_6][i].value
                field_type = ParseModelFieldType(field_origin_type)
                field_enum = ParseModelFieldEnum(field_origin_type)
                if field_enum:
                    gen_enums.append(field_enum)
                field_unique = ParseModelFieldUnique(sheet[ROW_7][i].value)
                field_required = sheet[ROW_8][i].value
                field_server = sheet[ROW_9][i].value
                field_client = sheet[ROW_10][i].value
                last_field_name = field_name
                # create field instance
                field = ModelField(
                    model,
                    field_type,
                    field_origin_type,
                    field_name,
                    field_description,
                    field_unique,
                    field_required,
                    field_server,
                    field_client)
                model.AddField(field)
            # save model to generate
            gen_models.append(model)

    # fill foreign keys in correct position
    for model in gen_models:
        foreign_key_fields = list()
        for field in model.fields:
            if field.type == ModelFieldType.ForeignKey:
                foreign_key_fields.append(field)
        # remove wrong position
        for field in foreign_key_fields:
            model.fields.remove(field)
        # add to correct position
        for field in foreign_key_fields:
            target_model_name = ParseForeignKeyName(field.origin_type)
            target_model = None
            for search_model in gen_models:
                if search_model.name == target_model_name:
                    target_model = search_model
                    break
            if not target_model:
                raise Exception('Cannot find model ' + target_model_name)
            if target_model.name != field.sourceModel.name and not field.foreignKey:
                field.foreignKey = field.name
                # update field infomation
                field.name = field.sourceModel.name
            target_model.AddField(field)

    return (gen_models, gen_enums)

def ParseData(xls_file, sheet_name):
    wb = load_workbook(xls_file, read_only=True)
    sheet = wb[sheet_name]
    data = list()
    # load field name
    field_names = list()
    for i in range(1, len(sheet[ROW_4])):
        field_names.append(sheet[ROW_4][i].value)
    data.append(field_names)
    # # load field type
    # field_types = list()
    # for i in range(1, len(sheet[ROW_6])):
    #     field_types.append(sheet[ROW_6][i].value)
    # data.append(field_types)
    # load data row
    for row in sheet.iter_rows(min_row=DATA_START_ROW):
        row_data = list()
        for i in range(1, len(row)):
            row_data.append(row[i].value)
        data.append(row_data)

    return data
