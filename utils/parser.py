#!/usr/bin/python
# -*- coding: UTF-8 -*-
import os
from openpyxl import load_workbook
from utils.model import Model, ModelField, ModelFieldType, ParseOneToOneFieldName, ParseForeignKeyName
from utils.enum import Enum, EnumField

COL_2_ROW_1 = 'B1'  # ClassName
COL_2_ROW_2 = 'B2'  # ClassDescription
COL_2_ROW_3 = 'B3'  # ClassGroup

ROW_4 = '4'         # FieldName
ROW_5 = '5'         # FieldDescription
ROW_6 = '6'         # Type
ROW_7 = '7'         # Uniqueness
ROW_8 = '8'         # Required
ROW_9 = '9'         # Server
ROW_10 = '10'       # Client

def ParseModelFieldType(field):
    if field == 'id':
        return ModelFieldType.IntegerField
    elif field == 'bool':
        return ModelFieldType.BooleanField
    elif field == 'string':
        return ModelFieldType.CharField
    elif field == 'int':
        return ModelFieldType.IntegerField
    elif field == 'float':
        return ModelFieldType.FloatField
    elif field.startswith('id('):
        return ModelFieldType.OneToOneField
    elif field.startswith('array(id('):
        return ModelFieldType.ForeignKey
    elif field.startswith('array('):
        return ModelFieldType.TextField
    elif field.startswith('enum('):
        return ModelFieldType.IntegerField
    elif field.startswith('Enum'):
        return ModelFieldType.IntegerField
    raise Exception('Unknown field type ' + field)

def ParseModelFieldUnique(field):
    if field == 'None':
        return False
    return True

def ParseModelFieldEnum(field):
    if field.startswith('enum('):
        content_parts = field[field.index('(') + 1:field.rindex(')')].split(',')
        enum_name = content_parts[0]
        enum = Enum(enum_name)
        # parse enum fields
        for i in range(1, len(content_parts)):
            content_part = content_parts[i]
            field_name = content_part[0:content_part.index('(')]
            field_description = content_part[content_part.index('(') + 1:content_part.index(')')]
            field = EnumField(field_name, i - 1, field_description)
            enum.AddField(field)
        return enum
    return None

def Parse(source):
    xls_files = list()
    print("Scanning xls file in", source)
    for file in os.listdir(source):
        if file.endswith(".xlsx") and not file.startswith("~"):
            xls_files.append(os.path.join(source, file))

    if len(xls_files) == 0:
        print("Could not find xls file in", source)
        return False
    print("Found xls files", xls_files)

    gen_models = list()
    gen_enums = list()

    for xls_file in xls_files:
        print("Parsing xls file", xls_file)
        wb = load_workbook(xls_file, read_only=True)
        print("Found xls sheets", wb.sheetnames)
        # parse xls sheet
        for sheet in wb:
            print('Parsing xls sheet', sheet.title)
            # load model base info
            class_name = sheet[COL_2_ROW_1].value
            class_description = sheet[COL_2_ROW_2].value
            class_group = sheet[COL_2_ROW_3].value
            # create model instance
            model = Model(
                class_name,
                class_description,
                class_group)
            # load fields base info
            last_field_name = ''
            for i in range(1, len(sheet[ROW_4])):
                field_name = sheet[ROW_4][i].value
                # array field type
                print("Parsing field", field_name)
                if field_name == last_field_name:
                    continue
                field_description = sheet[ROW_5][i].value
                field_origin_type = sheet[ROW_6][i].value
                field_type = ParseModelFieldType(field_origin_type)
                field_enum = ParseModelFieldEnum(field_origin_type)
                if field_enum:
                    gen_enums.append(field_enum)
                field_unique = ParseModelFieldUnique(sheet[ROW_7][i].value)
                field_required = sheet[ROW_8][i].value
                field_server = sheet[ROW_9][i].value
                field_client = sheet[ROW_10][i].value
                last_field_name = field_name
                # create field instance
                field = ModelField(
                    field_type,
                    field_origin_type,
                    field_name,
                    field_description,
                    field_unique,
                    field_required,
                    field_server,
                    field_client)
                model.AddField(field)
            # save model to generate
            gen_models.append(model)

    return (gen_models, gen_enums)